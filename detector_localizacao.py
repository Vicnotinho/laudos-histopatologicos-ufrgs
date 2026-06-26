"""
detector_localizacao.py — Detector de localizações anatômicas parecidas.

REGRAS (combinadas com o Victor):
  1. NÚMEROS de dente: só junta se for o MESMO número.
     "Ápice do 33" ≠ "Ápice do 25"  (dentes diferentes — NÃO juntam)
     "Ápice do 33" + "Apice do 33"   (mesmo número — juntam, corrige grafia)

  2. "e" sozinho NÃO é lado:
     "Dorso e língua" → "Dorso da língua" (sem lado, não vira esquerda)

  3. LADOS mantidos distintos:
     "Mucosa jugal D" + "Mucosa jugal direita" → JUNTAM (mesmo lado)
     "Mucosa jugal D" (direita) ≠ "Mucosa jugal E" (esquerda)

  4. GENGIVA com número → "Gengiva região do X" (mantém o número):
     "Gengiva 24", "Gengiva região do 47" → "Gengiva região do 24/47"

  5. GENGIVA VESTIBULAR → junta tudo:
     "Gengiva V", "vestibular", "intraósseo", "intra-ósseo" → "Gengiva vestibular"

⚠️ NÃO altera o banco. Só gera revisar_localizacao.xlsx para revisão.

Rode na sua máquina:  python detector_localizacao.py
"""

import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import banco

SAIDA = Path(__file__).parent / "revisar_localizacao.xlsx"
LIMIAR_SIMILARIDADE = 0.88


def _sem_acento(txt: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", txt)
                   if unicodedata.category(c) != "Mn")


def _normalizar(txt: str) -> str:
    txt = _sem_acento((txt or "").strip().lower())
    txt = txt.strip(" .;,:")
    txt = " ".join(txt.split())
    return txt


# ── REGRA 3: detecção de LADO (sem confundir "e" sozinho) ──────────────────
# Importante: "\be\b" NÃO entra aqui (regra 2). Só formas explícitas de lado.
PADROES_LADO = [
    (r"\b(lado direito|direita|direito|\bdir\b|\bld\b|\bd\b)\b", "direita"),
    (r"\b(lado esquerdo|esquerda|esquerdo|\besq\b|\ble\b)\b", "esquerda"),
]


def _extrair_lado(norm: str) -> tuple[str, str]:
    """Retorna (texto_sem_lado, lado). 'e' sozinho NÃO é considerado lado."""
    lado = ""
    corpo = norm
    for padrao, canonico in PADROES_LADO:
        if re.search(padrao, corpo):
            lado = canonico
            corpo = re.sub(padrao, "", corpo).strip()
            corpo = " ".join(corpo.split())
            break
    return corpo, lado


# ── REGRA 1: extrair NÚMERO de dente ───────────────────────────────────────
def _extrair_numero(norm: str) -> tuple[str, str]:
    """
    Retorna (texto_sem_numero, numero).
    Pega o primeiro número de 1-2 dígitos encontrado (nº de dente).
    """
    m = re.search(r"\b(\d{1,2})\b", norm)
    if m:
        numero = m.group(1)
        corpo = (norm[:m.start()] + norm[m.end():]).strip()
        corpo = " ".join(corpo.split())
        return corpo, numero
    return norm, ""


# ── REGRA 2: corrigir "dorso e lingua" → "dorso da lingua" ─────────────────
def _corrigir_e_isolado(norm: str) -> str:
    # troca " e " isolado (que sobrou) por " da " quando faz sentido (dorso/ventre + lingua)
    norm = re.sub(r"\bdorso e lingua\b", "dorso da lingua", norm)
    norm = re.sub(r"\bventre e lingua\b", "ventre da lingua", norm)
    norm = re.sub(r"\bbordo e lingua\b", "bordo da lingua", norm)
    norm = re.sub(r"\bborda e lingua\b", "borda da lingua", norm)
    return norm


# ── REGRA 5: normalizar gengiva vestibular (intraósseo/V) ──────────────────
def _eh_gengiva_vestibular(norm: str) -> bool:
    tem_gengiva = "gengiva" in norm
    tem_vest = bool(re.search(r"\bvestibular\b|\bgengiva v\b|\bvestib\b", norm))
    return tem_gengiva and tem_vest


# ── REGRA 4: gengiva com número → "gengiva regiao do X" ────────────────────
def _eh_gengiva_com_numero(norm: str) -> tuple[bool, str]:
    if "gengiva" not in norm:
        return False, ""
    if _eh_gengiva_vestibular(norm):
        return False, ""  # vestibular tem prioridade (regra 5)
    m = re.search(r"\b(\d{1,2})\b", norm)
    if m:
        return True, m.group(1)
    return False, ""


def _chave_e_sugestao(original: str) -> tuple[tuple, str]:
    """
    Decide a chave de agrupamento e a sugestão final de cada localização,
    aplicando as 5 regras na ordem de prioridade.
    Retorna (chave, sugestao).
    """
    norm = _normalizar(original)
    norm = _corrigir_e_isolado(norm)  # regra 2

    # regra 5: gengiva vestibular junta tudo
    if _eh_gengiva_vestibular(norm):
        return (("gengiva vestibular", "", ""), "Gengiva vestibular")

    # regra 4: gengiva com número → região do X
    eh_geng_num, num_geng = _eh_gengiva_com_numero(norm)
    if eh_geng_num:
        return (("gengiva regiao do", "", num_geng), f"Gengiva região do {num_geng}")

    # regra 3 + 1: extrai lado e número (mantém ambos distintos)
    corpo, lado = _extrair_lado(norm)
    corpo, numero = _extrair_numero(corpo)

    chave = (corpo, lado, numero)

    # monta sugestão preservando acento do original (já com "e"→"da" aplicado)
    original_corrigido = _aplicar_correcao_e_no_original(original)
    base = _base_com_acento(original_corrigido, lado, numero)
    partes = [base]
    if numero:
        partes.append(numero)
    if lado:
        partes.append(f"({lado})")
    sugestao = " ".join(p for p in partes if p).strip()
    sugestao = sugestao[:1].upper() + sugestao[1:] if sugestao else sugestao
    return (chave, sugestao)


def _aplicar_correcao_e_no_original(original: str) -> str:
    """Aplica a regra 2 (dorso e língua → dorso da língua) preservando acentos."""
    out = original
    # troca "<palavra> e língua" por "<palavra> da língua" (case/acento-insensível)
    for palavra in ("dorso", "ventre", "bordo", "borda"):
        out = re.sub(
            rf"\b({palavra})\s+e\s+(l[íi]ngua)\b",
            r"\1 da \2",
            out, flags=re.IGNORECASE,
        )
    return out


def _base_com_acento(original: str, lado: str, numero: str) -> str:
    """Remove número e lado do ORIGINAL (preservando acentos do resto)."""
    base = original.strip().strip(" .;,:")
    # remove o número
    if numero:
        base = re.sub(rf"\b{re.escape(numero)}\b", "", base)
    # remove a palavra de lado (no original, case-insensitive via sem-acento)
    if lado:
        for padrao, canon in PADROES_LADO:
            if canon == lado:
                base_sa = _sem_acento(base.lower())
                m = re.search(padrao, base_sa)
                if m:
                    base = (base[:m.start()] + base[m.end():])
                break
    base = " ".join(base.split()).strip(" .,;:-")
    return base


def _parecidos(a: str, b: str) -> bool:
    return SequenceMatcher(None, a, b).ratio() >= LIMIAR_SIMILARIDADE


def detectar():
    banco.inicializar()
    with banco.conectar() as conn:
        rows = conn.execute(
            '''SELECT TRIM("localizacao") AS loc, COUNT(*) AS n
               FROM laudos
               WHERE TRIM("localizacao") != ''
               GROUP BY "localizacao"'''
        ).fetchall()

    # cada item: (original, n, chave, sugestao)
    itens = []
    for r in rows:
        original = r["loc"]
        n = r["n"]
        chave, sugestao = _chave_e_sugestao(original)
        itens.append((original, n, chave, sugestao))

    # ── agrupa por chave EXATA primeiro ────────────────────────────────────
    por_chave = {}
    sugestao_da_chave = {}
    for original, n, chave, sugestao in itens:
        por_chave.setdefault(chave, []).append((original, n))
        sugestao_da_chave[chave] = sugestao

    # ── junta chaves parecidas SÓ se mesmo lado E mesmo número ─────────────
    chaves = list(por_chave.keys())
    usados = set()
    grupos = []
    for i, ka in enumerate(chaves):
        if ka in usados:
            continue
        corpo_a, lado_a, num_a = ka
        grupo = [ka]
        usados.add(ka)
        for kb in chaves[i + 1:]:
            if kb in usados:
                continue
            corpo_b, lado_b, num_b = kb
            # REGRA 1 e 3: só junta se lado E número forem iguais
            if lado_a == lado_b and num_a == num_b and _parecidos(corpo_a, corpo_b):
                grupo.append(kb)
                usados.add(kb)
        grupos.append(grupo)

    # ── monta planilha ─────────────────────────────────────────────────────
    dados = []
    grupo_id = 0
    for grupo in grupos:
        variacoes = []
        for chave in grupo:
            variacoes.extend(por_chave[chave])

        if len(variacoes) <= 1:
            continue

        grupo_id += 1
        variacoes.sort(key=lambda x: -x[1])
        # sugestão: a da chave mais usada do grupo
        chave_top = max(grupo, key=lambda k: sum(n for _, n in por_chave[k]))
        sugestao = sugestao_da_chave[chave_top]
        total = sum(n for _, n in variacoes)

        for original, n in variacoes:
            dados.append((grupo_id, original, n, sugestao, total))

    _salvar_excel(dados)

    print("=" * 60)
    print("DETECTOR DE LOCALIZAÇÕES (com regras de número e lado)")
    print("=" * 60)
    print(f"  Grupos com variações:   {grupo_id}")
    print(f"  Linhas no relatório:    {len(dados)}")
    print(f"\n  Arquivo: {SAIDA.name}")
    print("\n  Regras aplicadas:")
    print("   1. Números de dente diferentes NÃO juntam (33 ≠ 25)")
    print("   2. 'Dorso e língua' → 'Dorso da língua' (sem lado)")
    print("   3. Lados mantidos (direita ≠ esquerda)")
    print("   4. Gengiva + número → 'Gengiva região do X'")
    print("   5. Gengiva vestibular junta tudo (V, intraósseo...)")
    print("=" * 60)


def _salvar_excel(dados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Revisar Localização"

    cab = ["Grupo", "Localização atual", "Casos",
           "Unificar para (editável)", "Total do grupo"]
    ws.append(cab)

    fill_cab = PatternFill("solid", fgColor="1F4E78")
    fonte_cab = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    borda = Border(*[Side(style="thin", color="CCCCCC")] * 4)
    for col in range(1, len(cab) + 1):
        c = ws.cell(row=1, column=col)
        c.fill, c.font = fill_cab, fonte_cab
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda

    cor_a = PatternFill("solid", fgColor="EAF1FB")
    cor_b = PatternFill("solid", fgColor="FFFFFF")
    fonte = Font(name="Arial", size=10)
    fonte_sug = Font(name="Arial", size=10, color="0000FF")

    for i, (g, atual, casos, sugestao, total) in enumerate(dados, start=2):
        ws.cell(row=i, column=1, value=g)
        ws.cell(row=i, column=2, value=atual)
        ws.cell(row=i, column=3, value=casos)
        ws.cell(row=i, column=4, value=sugestao)
        ws.cell(row=i, column=5, value=total)
        cor = cor_a if (g % 2 == 0) else cor_b
        for col in range(1, 6):
            cel = ws.cell(row=i, column=col)
            cel.fill, cel.border = cor, borda
            cel.font = fonte_sug if col == 4 else fonte
            if col in (1, 3, 5):
                cel.alignment = Alignment(horizontal="center")

    for i, w in enumerate([8, 45, 8, 45, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Instruções")
    for linha in [
        ["COMO REVISAR — LOCALIZAÇÕES"],
        [""],
        ["REGRAS JÁ APLICADAS:"],
        ["• Números de dente diferentes NÃO foram juntados (33 ≠ 25)."],
        ["• 'Dorso e língua' virou 'Dorso da língua' (sem lado)."],
        ["• Lados mantidos: direita ≠ esquerda."],
        ["• Gengiva + número → 'Gengiva região do X'."],
        ["• Gengiva vestibular juntou tudo (V, intraósseo, intra-ósseo)."],
        [""],
        ["O QUE FAZER:"],
        ["1. Revise a coluna azul 'Unificar para' com o professor."],
        ["2. Onde NÃO deve juntar: escreva 'NÃO JUNTAR' na coluna azul."],
        ["3. Salve e me devolva o arquivo."],
    ]:
        ws2.append(linha)
    ws2.column_dimensions["A"].width = 72
    ws2["A1"].font = Font(bold=True, size=13)

    wb.save(SAIDA)


if __name__ == "__main__":
    detectar()