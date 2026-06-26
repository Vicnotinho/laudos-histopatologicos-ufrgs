"""
detector_diagnosticos_clinico.py — Detector do DIAGNÓSTICO CLÍNICO.

Igual ao detector_diagnosticos_v2.py, mas analisa o campo
"Diagnóstico Clínico" (a hipótese do dentista) em vez do histopatológico.
Serve para permitir as mesmas estatísticas com o diagnóstico clínico.

  • Remove prefixos como "Sugestivo de", "Compatível com" para agrupar.
  • Junta erros de digitação automaticamente (sugere a forma mais usada).
  • Mostra quais prefixos foram encontrados, para você conferir.

⚠️ NÃO altera o banco. Só gera revisar_diagnosticos_clinico.xlsx para revisão.

Rode na sua máquina:  python detector_diagnosticos_clinico.py
"""

import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import banco

SAIDA = Path(__file__).parent / "revisar_diagnosticos_clinico.xlsx"
LIMIAR_SIMILARIDADE = 0.88

# Prefixos a remover para agrupar (em ordem: mais específicos primeiro).
# São testados sobre o texto SEM acento e minúsculo.
PREFIXOS = [
    "sugestivo de", "sugestiva de", "sugere", "sugestao de", "sugestao",
    "compativel com o", "compativel com a", "compativel com",
    "compativel a", "compativel",
    "achados compativeis com", "achados de", "achado de",
    "aspectos de", "aspecto de",
    "quadro compativel com", "quadro de",
    "diagnostico de", "diagnostico compativel com",
]


def _sem_acento(txt: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", txt)
                   if unicodedata.category(c) != "Mn")


def _normalizar(txt: str) -> str:
    """minúsculo, sem acento, sem pontuação nas pontas, espaços únicos."""
    txt = _sem_acento((txt or "").strip().lower())
    txt = txt.strip(" .;,:")
    txt = " ".join(txt.split())
    return txt


def _remover_prefixo(norm: str) -> tuple[str, str]:
    """
    Retorna (texto_sem_prefixo, prefixo_encontrado).
    Se não houver prefixo, devolve (norm, '').
    """
    for p in PREFIXOS:
        if norm.startswith(p + " "):
            return norm[len(p):].strip(), p
    return norm, ""


def _parecidos(a: str, b: str) -> bool:
    return SequenceMatcher(None, a, b).ratio() >= LIMIAR_SIMILARIDADE


def detectar():
    banco.inicializar()
    with banco.conectar() as conn:
        rows = conn.execute(
            '''SELECT TRIM("diagnostico_clinico") AS diag, COUNT(*) AS n
               FROM laudos
               WHERE TRIM("diagnostico_clinico") != ''
               GROUP BY "diagnostico_clinico"'''
        ).fetchall()

    # cada item: original, contagem, núcleo (sem prefixo), prefixo
    itens = []
    prefixos_usados = {}
    for r in rows:
        original = r["diag"]
        n = r["n"]
        norm = _normalizar(original)
        nucleo, prefixo = _remover_prefixo(norm)
        itens.append((original, n, nucleo, prefixo))
        if prefixo:
            prefixos_usados[prefixo] = prefixos_usados.get(prefixo, 0) + n

    # ── agrupa pelo NÚCLEO (sem prefixo) ───────────────────────────────────
    por_nucleo = {}
    for original, n, nucleo, prefixo in itens:
        por_nucleo.setdefault(nucleo, []).append((original, n))

    # ── junta núcleos PARECIDOS (erros de digitação) ───────────────────────
    nucleos = list(por_nucleo.keys())
    usados = set()
    grupos = []
    for i, na in enumerate(nucleos):
        if na in usados:
            continue
        grupo = [na]
        usados.add(na)
        for nb in nucleos[i + 1:]:
            if nb in usados:
                continue
            if _parecidos(na, nb):
                grupo.append(nb)
                usados.add(nb)
        grupos.append(grupo)

    # ── monta a planilha ───────────────────────────────────────────────────
    dados = []
    grupo_id = 0
    for grupo in grupos:
        variacoes = []
        for nuc in grupo:
            variacoes.extend(por_nucleo[nuc])

        if len(variacoes) <= 1:
            continue

        grupo_id += 1
        variacoes.sort(key=lambda x: -x[1])
        # sugestão = a forma original mais usada, mas sem o prefixo e com inicial maiúscula
        original_top = variacoes[0][0]
        norm_top = _normalizar(original_top)
        nucleo_top, prefixo_top = _remover_prefixo(norm_top)

        if prefixo_top:
            # tinha prefixo: usa o núcleo (sem acento), capitalizado
            base = nucleo_top
        else:
            # não tinha prefixo: preserva o ORIGINAL (com acento), só ajusta inicial
            base = original_top.strip().strip(" .;,:")
            # se a forma mais usada tinha prefixo mas outra não, prefira a sem prefixo
            for orig, _ in variacoes:
                n_orig = _normalizar(orig)
                nuc, pref = _remover_prefixo(n_orig)
                if not pref:
                    base = orig.strip().strip(" .;,:")
                    break

        sugestao = base[:1].upper() + base[1:] if base else base
        total = sum(n for _, n in variacoes)

        for original, n in variacoes:
            dados.append((grupo_id, original, n, sugestao, total))

    _salvar_excel(dados, prefixos_usados)

    print("=" * 60)
    print("DETECTOR v2 — com remoção de prefixos")
    print("=" * 60)
    print(f"  Grupos com variações:   {grupo_id}")
    print(f"  Linhas no relatório:    {len(dados)}")
    print(f"\n  Prefixos encontrados ({len(prefixos_usados)}):")
    for p, n in sorted(prefixos_usados.items(), key=lambda x: -x[1]):
        print(f"     '{p}'  ({n} casos)")
    print(f"\n  Arquivo: {SAIDA.name}")
    print("=" * 60)


def _salvar_excel(dados, prefixos_usados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Revisar Diagnóstico Clínico"

    cab = ["Grupo", "Diagnóstico atual", "Casos",
           "Unificar para (editável)", "Total do grupo"]
    ws.append(cab)

    fill_cab = PatternFill("solid", fgColor="1F4E78")
    fonte_cab = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    borda = Border(*[Side(style="thin", color="CCCCCC")] * 4)
    for col in range(1, len(cab) + 1):
        c = ws.cell(row=1, column=col)
        c.fill, c.font = fill_cab, fonte_cab
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda

    cor_a = PatternFill("solid", fgColor="EAF1FB")
    cor_b = PatternFill("solid", fgColor="FFFFFF")
    fonte = Font(name="Arial", size=10)
    fonte_sug = Font(name="Arial", size=10, color="0000FF")

    for i, (g, atual, casos, sugestao, total) in enumerate(dados, start=2):
        ws.cell(row=i, column=1, value=g)
        ws.cell(row=i, column=2, value=atual)
        ws.cell(row=i, column=3, value=casos)
        ws.cell(row=i, column=4, value=sugestao)
        ws.cell(row=i, column=5, value=total)
        cor = cor_a if (g % 2 == 0) else cor_b
        for col in range(1, 6):
            cel = ws.cell(row=i, column=col)
            cel.fill, cel.border = cor, borda
            cel.font = fonte_sug if col == 4 else fonte
            if col in (1, 3, 5):
                cel.alignment = Alignment(horizontal="center")

    for i, w in enumerate([8, 50, 8, 50, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # aba de prefixos encontrados
    ws3 = wb.create_sheet("Prefixos removidos")
    ws3.append(["Prefixo encontrado", "Casos"])
    for c in ("A1", "B1"):
        ws3[c].font = Font(bold=True)
    for p, n in sorted(prefixos_usados.items(), key=lambda x: -x[1]):
        ws3.append([p, n])
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 10

    # aba de instruções
    ws2 = wb.create_sheet("Instruções")
    for linha in [
        ["COMO REVISAR"],
        [""],
        ["• Prefixos como 'Sugestivo de' e 'Compatível com' já foram"],
        ["  removidos para agrupar (ex: 'Sugestivo de rânula' = 'Rânula')."],
        ["• Erros de digitação já foram agrupados automaticamente."],
        ["• Coluna azul = sugestão (forma mais usada, sem prefixo)."],
        [""],
        ["O QUE FAZER:"],
        ["1. Confira a coluna azul com o professor."],
        ["2. Onde NÃO deve juntar (doenças diferentes, ex: odontoma"],
        ["   complexo vs composto): escreva 'NÃO JUNTAR' na coluna azul."],
        ["3. Salve e me devolva o arquivo."],
        [""],
        ["Veja também a aba 'Prefixos removidos' para conferir."],
    ]:
        ws2.append(linha)
    ws2.column_dimensions["A"].width = 70
    ws2["A1"].font = Font(bold=True, size=13)

    wb.save(SAIDA)


if __name__ == "__main__":
    detectar()