"""
detector_diagnosticos.py — Acha diagnósticos PARECIDOS para você revisar.

⚠️ Este script NÃO altera nada. Ele só gera um relatório (revisar_diagnosticos.xlsx)
   com grupos de diagnósticos parecidos, para você levar ao professor e decidir
   quais devem virar um só.

Como funciona:
   • Agrupa diagnósticos cujo texto normalizado (sem acento, sem maiúscula,
     sem pontuação) é quase igual.
   • Mostra quantos casos cada variação tem.
   • Gera um Excel já formatado (abre certo, com colunas separadas).

Rode na sua máquina:  python detector_diagnosticos.py
Depois abra:          revisar_diagnosticos.xlsx
"""

import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import banco

SAIDA = Path(__file__).parent / "revisar_diagnosticos.xlsx"

# quão parecidos precisam ser para entrarem no mesmo grupo (0 a 1)
LIMIAR_SIMILARIDADE = 0.86


def _normalizar(txt: str) -> str:
    txt = (txt or "").strip().lower()
    txt = "".join(c for c in unicodedata.normalize("NFD", txt)
                  if unicodedata.category(c) != "Mn")
    txt = txt.rstrip(".").strip()
    # remove espaços duplicados
    txt = " ".join(txt.split())
    return txt


def _parecidos(a: str, b: str) -> bool:
    return SequenceMatcher(None, a, b).ratio() >= LIMIAR_SIMILARIDADE


def detectar():
    banco.inicializar()
    with banco.conectar() as conn:
        rows = conn.execute(
            '''SELECT TRIM("diagnostico_histopatologico") AS diag, COUNT(*) AS n
               FROM laudos
               WHERE TRIM("diagnostico_histopatologico") != ''
               GROUP BY "diagnostico_histopatologico"'''
        ).fetchall()

    # lista de (texto_original, contagem, texto_normalizado)
    itens = [(r["diag"], r["n"], _normalizar(r["diag"])) for r in rows]

    # ── 1) agrupa exatamente iguais após normalização ──────────────────────
    por_norm = {}
    for original, n, norm in itens:
        por_norm.setdefault(norm, []).append((original, n))

    # ── 2) entre os grupos normalizados, junta os PARECIDOS ────────────────
    normas = list(por_norm.keys())
    usados = set()
    grupos = []

    for i, na in enumerate(normas):
        if na in usados:
            continue
        grupo = [na]
        usados.add(na)
        for nb in normas[i + 1:]:
            if nb in usados:
                continue
            if _parecidos(na, nb):
                grupo.append(nb)
                usados.add(nb)
        grupos.append(grupo)

    # ── monta os dados, alternando cor por grupo ───────────────────────────
    dados = []  # (grupo, atual, casos, sugestao, total)
    grupo_id = 0
    for grupo in grupos:
        variacoes = []
        for norm in grupo:
            variacoes.extend(por_norm[norm])

        if len(variacoes) <= 1:
            continue

        grupo_id += 1
        variacoes.sort(key=lambda x: -x[1])
        sugestao = variacoes[0][0]
        total = sum(n for _, n in variacoes)

        for original, n in variacoes:
            dados.append((grupo_id, original, n, sugestao, total))

    # ── gera o Excel formatado ─────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Revisar Diagnósticos"

    cabecalhos = ["Grupo", "Diagnóstico atual", "Casos",
                  "Unificar para (editável)", "Total do grupo"]
    ws.append(cabecalhos)

    # estilo do cabeçalho
    fill_cab = PatternFill("solid", fgColor="1F4E78")
    fonte_cab = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    borda = Border(*[Side(style="thin", color="CCCCCC")] * 4)
    for col in range(1, len(cabecalhos) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill_cab
        c.font = fonte_cab
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda

    # cores alternadas por grupo (facilita enxergar onde começa/termina)
    cor_a = PatternFill("solid", fgColor="EAF1FB")
    cor_b = PatternFill("solid", fgColor="FFFFFF")
    fonte = Font(name="Arial", size=10)
    fonte_sug = Font(name="Arial", size=10, color="0000FF")  # azul = editável

    for i, (g, atual, casos, sugestao, total) in enumerate(dados, start=2):
        ws.cell(row=i, column=1, value=g)
        ws.cell(row=i, column=2, value=atual)
        ws.cell(row=i, column=3, value=casos)
        ws.cell(row=i, column=4, value=sugestao)
        ws.cell(row=i, column=5, value=total)

        cor = cor_a if (g % 2 == 0) else cor_b
        for col in range(1, 6):
            cel = ws.cell(row=i, column=col)
            cel.fill = cor
            cel.border = borda
            cel.font = fonte_sug if col == 4 else fonte
            if col in (1, 3, 5):
                cel.alignment = Alignment(horizontal="center")

    # larguras
    larguras = [8, 48, 8, 48, 14]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # congela cabeçalho
    ws.freeze_panes = "A2"

    # aba de instruções
    ws2 = wb.create_sheet("Instruções")
    instrucoes = [
        ["COMO REVISAR ESTE ARQUIVO"],
        [""],
        ["1. Cada GRUPO (cores alternadas) reúne diagnósticos parecidos."],
        ["2. A coluna azul 'Unificar para' tem a sugestão (forma mais usada)."],
        ["3. Revise com o professor:"],
        ["   • Se o grupo DEVE virar um só: deixe/ajuste o nome na coluna azul."],
        ["   • Se NÃO deve juntar (são diferentes): apague o nome da coluna azul"],
        ["     nas linhas que não entram, ou escreva 'NÃO JUNTAR'."],
        ["4. Salve e me devolva este arquivo que eu aplico no banco."],
        [""],
        ["Dica: a coluna 'Casos' mostra quantos laudos têm aquela grafia."],
        ["      A grafia com mais casos costuma ser a 'correta'."],
    ]
    for linha in instrucoes:
        ws2.append(linha)
    ws2.column_dimensions["A"].width = 70
    ws2["A1"].font = Font(name="Arial", bold=True, size=13)

    wb.save(SAIDA)

    print("=" * 60)
    print("DETECÇÃO DE DIAGNÓSTICOS PARECIDOS")
    print("=" * 60)
    print(f"  Grupos com variações encontrados: {grupo_id}")
    print(f"  Linhas no relatório:              {len(dados)}")
    print(f"\n  Relatório salvo em: {SAIDA.name}")
    print("\n  PRÓXIMO PASSO:")
    print("  1. Abra o Excel (já vem com colunas separadas e cores).")
    print("  2. Revise com o professor na coluna azul 'Unificar para'.")
    print("  3. Me devolva o arquivo revisado que eu aplico no banco.")
    print("=" * 60)


if __name__ == "__main__":
    detectar()